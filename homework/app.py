import openpyxl as xl


wb = xl.load_workbook("test1.xlsx")
sheet1 = wb["Sheet1"]
# cell1 = sheet1["a1"]
cell1 = sheet1.cell(1, 2)

for row in range(2, sheet1.max_row + 1):
    price = sheet1.cell(row, 3).value
    correct_price = price * 0.9
    sheet1.cell(row, 4).value = correct_price
wb.save("test2.xlsx")
